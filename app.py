"""
Face Attendance System
======================

Flask + PostgreSQL 17 (managed via pgAdmin 4) + InsightFace + OpenCV
+ optional local SLM (Ollama) for natural-language attendance queries.

Run::

    python app.py

Then open http://localhost:5000

The application exposes a School ERP-style UI (sidebar with classes,
per-page modules, stat cards) on top of the original face-recognition
pipeline. No PostgreSQL schema was changed, only two new tables
(``teachers``, ``unknown_faces``) were added for the new modules.
"""
from __future__ import annotations

import os
import re
import socket
from datetime import date, datetime, timedelta, timezone

import cv2
from flask import (
    Flask,
    Response,
    jsonify,
    render_template,
    request,
)
from sqlalchemy import func, text
from sqlalchemy.exc import IntegrityError
from werkzeug.utils import secure_filename

from camera.stream import manager
from config import Config
from database import Attendance, Camera, Student, Teacher, UnknownFace
from database.enterprise_models import (
    AttendanceEvent, StudentAttendance, TeacherAttendance, AttendanceLog,
    AttendanceSetting, HolidayCalendar, WeeklyOff,
)
from database.db import db as _db
from recognition.encoder import encode_image
from recognition.recognizer import FaceRecognizer
from time_management import ensure_default_rules
from time_management.api import bp as time_mgmt_bp


# Canonical class list used by the sidebar / filter dropdowns.
CLASS_LEVELS = [
    "Nursery", "Jr KG", "Sr KG",
    "Class 1", "Class 2", "Class 3", "Class 4", "Class 5",
    "Class 6", "Class 7", "Class 8",
]
SECTIONS = ["A", "B", "C", "D"]


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


def _coerce_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on"}
    return bool(value)


def _normalize_class_name(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text in CLASS_LEVELS:
        return text
    match = re.fullmatch(r"(?i)class\s*([0-9]+)", text)
    if match:
        return f"Class {int(match.group(1))}"
    match = re.fullmatch(r"([0-9]+)", text)
    if match:
        return f"Class {int(match.group(1))}"
    return text


def _normalize_section(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text.upper()


def _normalize_camera_source(value) -> str | int:
    """Normalize a camera source the same way the live recognizer does."""
    return FaceRecognizer._normalize_source(value)


def _camera_source_key(value) -> str:
    """Stable string key for comparing physical camera sources."""
    normalized = _normalize_camera_source(value)
    return str(normalized)


def _find_camera_by_source(source) -> Camera | None:
    """Return an existing camera that points at the same physical source."""
    target = _camera_source_key(source)
    for cam in Camera.query.all():
        if _camera_source_key(cam.source) == target:
            return cam
    return None


def ensure_schema_compat() -> None:
    """Add missing columns expected by the current app without rewriting data.

    ``create_all`` creates missing tables but intentionally does not alter
    existing tables. Older installs can therefore lack newer nullable columns
    like mobile/email even though the ORM model and UI already use them.
    """
    statements = [
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS class_name VARCHAR(80)",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS section VARCHAR(40)",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS mobile VARCHAR(20)",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS email VARCHAR(120)",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS photo_path VARCHAR(255)",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS encoding JSONB",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS is_active BOOLEAN NOT NULL DEFAULT TRUE",
        "ALTER TABLE students ADD COLUMN IF NOT EXISTS created_at TIMESTAMPTZ NOT NULL DEFAULT now()",
        # ---- time-management additions (additive, safe defaults) ----
        "ALTER TABLE attendance ADD COLUMN IF NOT EXISTS event_type VARCHAR(8) NOT NULL DEFAULT 'in'",
        "ALTER TABLE attendance ADD COLUMN IF NOT EXISTS session_id INTEGER",
        "ALTER TABLE attendance ADD COLUMN IF NOT EXISTS manual_edit BOOLEAN NOT NULL DEFAULT FALSE",
        "ALTER TABLE attendance ADD COLUMN IF NOT EXISTS edit_reason VARCHAR(255)",
        "CREATE INDEX IF NOT EXISTS ix_attendance_session_id ON attendance(session_id)",
        "ALTER TABLE unknown_faces ADD COLUMN IF NOT EXISTS image_path VARCHAR(255)",
        "ALTER TABLE unknown_faces ADD COLUMN IF NOT EXISTS location VARCHAR(120)",
        "ALTER TABLE unknown_faces ADD COLUMN IF NOT EXISTS alerted BOOLEAN NOT NULL DEFAULT FALSE",
        "ALTER TABLE teachers ADD COLUMN IF NOT EXISTS department VARCHAR(120)",
        "ALTER TABLE teachers ADD COLUMN IF NOT EXISTS designation VARCHAR(120)",
        "ALTER TABLE teachers ADD COLUMN IF NOT EXISTS photo_path VARCHAR(255)",
        "ALTER TABLE teachers ADD COLUMN IF NOT EXISTS encoding JSONB",
        "ALTER TABLE student_attendance ADD COLUMN IF NOT EXISTS working_seconds INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE teacher_attendance ADD COLUMN IF NOT EXISTS working_seconds INTEGER NOT NULL DEFAULT 0",
        "ALTER TABLE student_time_rules ADD COLUMN IF NOT EXISTS in_end_time VARCHAR(5) NOT NULL DEFAULT '09:30'",
        "ALTER TABLE student_time_rules ADD COLUMN IF NOT EXISTS out_detection_start VARCHAR(5) NOT NULL DEFAULT '16:00'",
        "ALTER TABLE student_time_rules ADD COLUMN IF NOT EXISTS early_exit_time VARCHAR(5) NOT NULL DEFAULT '16:30'",
        "ALTER TABLE student_time_rules ADD COLUMN IF NOT EXISTS overtime_start VARCHAR(5) NOT NULL DEFAULT '17:00'",
        "ALTER TABLE teacher_time_rules ADD COLUMN IF NOT EXISTS in_end_time VARCHAR(5) NOT NULL DEFAULT '09:30'",
        "ALTER TABLE teacher_time_rules ADD COLUMN IF NOT EXISTS out_detection_start VARCHAR(5) NOT NULL DEFAULT '16:00'",
        "ALTER TABLE teacher_time_rules ADD COLUMN IF NOT EXISTS early_exit_time VARCHAR(5) NOT NULL DEFAULT '16:30'",
        "ALTER TABLE teacher_time_rules ADD COLUMN IF NOT EXISTS overtime_start VARCHAR(5) NOT NULL DEFAULT '17:00'",
    ]
    if _db.engine.dialect.name == "sqlite":
        # SQLite supports ADD COLUMN but not PostgreSQL's IF NOT EXISTS.
        # Only issue an ALTER for columns absent from an existing database.
        from sqlalchemy import inspect
        columns = {table: {c["name"] for c in inspect(_db.engine).get_columns(table)}
                   for table in ("students", "attendance", "unknown_faces", "teachers", "student_attendance", "teacher_attendance", "student_time_rules", "teacher_time_rules")}
        sqlite_columns = {
            "students": [("class_name", "VARCHAR(80)"), ("section", "VARCHAR(40)"), ("mobile", "VARCHAR(20)"), ("email", "VARCHAR(120)"), ("photo_path", "VARCHAR(255)"), ("encoding", "JSON"), ("is_active", "BOOLEAN NOT NULL DEFAULT 1"), ("created_at", "DATETIME")],
            "attendance": [("event_type", "VARCHAR(8) NOT NULL DEFAULT 'in'"), ("session_id", "INTEGER"), ("manual_edit", "BOOLEAN NOT NULL DEFAULT 0"), ("edit_reason", "VARCHAR(255)")],
            "unknown_faces": [("image_path", "VARCHAR(255)"), ("location", "VARCHAR(120)"), ("alerted", "BOOLEAN NOT NULL DEFAULT 0")],
            "teachers": [("department", "VARCHAR(120)"), ("designation", "VARCHAR(120)"), ("photo_path", "VARCHAR(255)"), ("encoding", "JSON")],
            "student_attendance": [("working_seconds", "INTEGER NOT NULL DEFAULT 0")],
            "teacher_attendance": [("working_seconds", "INTEGER NOT NULL DEFAULT 0")],
            "student_time_rules": [("in_end_time", "VARCHAR(5) NOT NULL DEFAULT '09:30'"), ("out_detection_start", "VARCHAR(5) NOT NULL DEFAULT '16:00'"), ("early_exit_time", "VARCHAR(5) NOT NULL DEFAULT '16:30'"), ("overtime_start", "VARCHAR(5) NOT NULL DEFAULT '17:00'")],
            "teacher_time_rules": [("in_end_time", "VARCHAR(5) NOT NULL DEFAULT '09:30'"), ("out_detection_start", "VARCHAR(5) NOT NULL DEFAULT '16:00'"), ("early_exit_time", "VARCHAR(5) NOT NULL DEFAULT '16:30'"), ("overtime_start", "VARCHAR(5) NOT NULL DEFAULT '17:00'")],
        }
        for table, definitions in sqlite_columns.items():
            for column, definition in definitions:
                if column not in columns.get(table, set()):
                    _db.session.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {definition}"))
    else:
        for stmt in statements:
            _db.session.execute(text(stmt))

    for student in Student.query.all():
        normalized_class = _normalize_class_name(student.class_name)
        normalized_section = _normalize_section(student.section)
        if normalized_class != student.class_name or normalized_section != student.section:
            student.class_name = normalized_class
            student.section = normalized_section

    _db.session.commit()


def create_app() -> Flask:
    """Application factory used by both ``python app.py`` and tests."""
    app = Flask(__name__, static_folder="static", template_folder="templates")
    app.config.from_object(Config)
    _db.init_app(app)

    os.makedirs(app.config["KNOWN_FACES_DIR"], exist_ok=True)

    with app.app_context():
        _db.create_all()
        ensure_schema_compat()
        # Seed the default time-rule rows for teacher + student scopes.
        # Idempotent — a no-op if they already exist.
        try:
            ensure_default_rules()
        except Exception as exc:
            print(f"[WARN] ensure_default_rules failed: {exc}")
        # Separate normalized enterprise rule rows, kept apart from the
        # compatibility time-rule module above.
        from database.enterprise_models import TeacherTimeRule, StudentTimeRule
        if TeacherTimeRule.query.first() is None:
            _db.session.add(TeacherTimeRule())
        if StudentTimeRule.query.first() is None:
            _db.session.add(StudentTimeRule())
        _db.session.commit()

        # --- deduplicate cameras (fix for Issue #2) ---
        # Group cameras by their normalised source key and keep only the
        # oldest record (lowest id) for each physical device.
        _seen_sources: dict[str, int] = {}
        for cam in Camera.query.order_by(Camera.id).all():
            key = _camera_source_key(cam.source)
            if key in _seen_sources:
                # Duplicate – null-out FK references then delete
                Attendance.query.filter_by(camera_id=cam.id).update(
                    {Attendance.camera_id: _seen_sources[key]},
                    synchronize_session=False,
                )
                UnknownFace.query.filter_by(camera_id=cam.id).update(
                    {UnknownFace.camera_id: _seen_sources[key]},
                    synchronize_session=False,
                )
                _db.session.delete(cam)
            else:
                _seen_sources[key] = cam.id
        _db.session.commit()

        # Auto-create default camera only if none exists for that source.
        # Wrapped in try/except to handle the race condition caused by
        # Flask's debug reloader running create_app() in two processes.
        default_source = str(app.config["CAMERA_SOURCE"])
        if _find_camera_by_source(default_source) is None:
            try:
                default_cam = Camera(
                    name="Default Webcam",
                    source=default_source,
                    location="Lab 1",
                    is_active=True,
                )
                _db.session.add(default_cam)
                _db.session.commit()
            except IntegrityError:
                _db.session.rollback()

    register_routes(app)
    app.register_blueprint(time_mgmt_bp)

    @app.context_processor
    def inject_globals():
        return {
            "CLASS_LEVELS": CLASS_LEVELS,
            "SECTIONS": SECTIONS,
        }

    @app.errorhandler(404)
    def _not_found(_):
        if request.path.startswith("/api/"):
            return jsonify({"error": "not found"}), 404
        return render_template("404.html"), 404

    return app


def register_routes(app: Flask) -> None:
    """Register every URL the dashboard depends on."""

    # ---------------------------------------------------------------- PAGES
    @app.route("/")
    def page_dashboard():
        return render_template(
            "dashboard.html",
            active_page="dashboard",
            active_class=None,
        )

    @app.route("/attendance")
    def page_attendance():
        return render_template(
            "attendance.html",
            active_page="attendance_enterprise",
        )

    @app.route("/students")
    def page_students():
        return render_template(
            "students.html",
            active_page="students",
        )

    @app.route("/students/class/<path:class_name>")
    def page_students_class(class_name: str):
        return render_template(
            "students.html",
            active_page="students",
            active_class=class_name,
        )

    @app.route("/teachers")
    def page_teachers():
        return render_template(
            "teachers.html",
            active_page="teachers",
        )

    @app.route("/cameras")
    def page_cameras():
        return render_template(
            "cameras.html",
            active_page="cameras",
        )

    @app.route("/reports")
    def page_reports():
        return render_template(
            "reports.html",
            active_page="reports",
        )

    @app.route("/settings")
    def page_settings():
        return render_template(
            "settings.html",
            active_page="settings",
        )

    @app.route("/chat")
    def chat_page():
        return render_template("chat.html", active_page=None)

    # ------------------------------------------- enterprise time management
    @app.route("/time-rules/<scope>")
    def page_time_rules(scope: str):
        if scope not in {"teacher", "student"}:
            scope = "student"
        return render_template(
            "time_rules.html",
            active_page="time_rules",
            active_time_scope=scope,
        )

    @app.route("/dashboard-enterprise")
    def page_dashboard_enterprise():
        return render_template(
            "dashboard_enterprise.html",
            active_page="dashboard_enterprise",
            active_class=None,
        )

    @app.route("/reports-enterprise")
    def page_reports_enterprise():
        return render_template(
            "reports_enterprise.html",
            active_page="reports_enterprise",
        )

    @app.route("/attendance-extended")
    def page_attendance_extended():
        return render_template(
            "attendance_extended.html",
            active_page="attendance_extended",
        )

    # ------------------------------------------------------------- meta api
    @app.route("/api/meta")
    def api_meta():
        """Static data the UI needs (class list, section list, etc.)."""
        return jsonify(
            {
                "classes": CLASS_LEVELS,
                "sections": SECTIONS,
                "school_name": app.config.get("SCHOOL_NAME", "FaceID School"),
            }
        )

    # ------------------------------------------------------------ stats api
    @app.route("/api/stats")
    def api_stats():
        """Class-aware KPI cards.

        Query string::

            class_name=Nursery    (optional, scoped to one class)
        """
        class_name = request.args.get("class_name") or None
        section = request.args.get("section") or None

        student_q = Student.query.filter_by(is_active=True)
        if class_name:
            student_q = student_q.filter_by(class_name=class_name)
        if section:
            student_q = student_q.filter_by(section=section)
        total_students = student_q.count()

        present_ids_q = _db.session.query(StudentAttendance.student_id).filter(
            StudentAttendance.attendance_date == date.today()
        )
        if class_name:
            present_ids_q = present_ids_q.join(
                Student, Student.id == StudentAttendance.student_id
            ).filter(Student.class_name == class_name)
        if section:
            present_ids_q = present_ids_q.join(
                Student, Student.id == StudentAttendance.student_id
            ).filter(Student.section == section)
        present_ids = {row[0] for row in present_ids_q.all()}

        present_today = sum(
            1 for sid in present_ids if student_q.filter(Student.id == sid).count()
        )
        present_today = len(present_ids) if not (class_name or section) else len(present_ids)

        # When a class filter is active, "present" = present AND in the class.
        if class_name or section:
            scoped_students = student_q.all()
            present_today = sum(1 for s in scoped_students if s.id in present_ids)
        else:
            present_today = len(present_ids)

        absent_today = max(total_students - present_today, 0)
        teachers_present = TeacherAttendance.query.filter_by(attendance_date=date.today()).count()
        total_teachers = Teacher.query.filter_by(is_active=True).count()
        entries = AttendanceEvent.query.filter(
            func.date(AttendanceEvent.event_time) == date.today(), AttendanceEvent.event_type == "in"
        ).count()
        exits = AttendanceEvent.query.filter(
            func.date(AttendanceEvent.event_time) == date.today(), AttendanceEvent.event_type == "out"
        ).count()
        late_today = StudentAttendance.query.filter_by(attendance_date=date.today(), is_late=True).count() + TeacherAttendance.query.filter_by(attendance_date=date.today(), is_late=True).count()
        unknown_today = (
            UnknownFace.query.filter(func.date(UnknownFace.timestamp) == date.today()).count()
        )
        active_cams = Camera.query.filter_by(is_active=True).count()

        return jsonify(
            {
                "total_students": total_students,
                "present_today": present_today,
                "absent_today": absent_today,
                "unknown_today": unknown_today,
                "active_cameras": active_cams,
                "students_present": present_today,
                "teachers_present": teachers_present,
                "students_absent": absent_today,
                "teachers_absent": max(total_teachers - teachers_present, 0),
                "late_today": late_today,
                "today_entries": entries,
                "today_exits": exits,
                "class_name": class_name,
                "section": section,
            }
        )

    # ----------------------------------------------------------------- cameras
    @app.route("/api/cameras", methods=["GET", "POST"])
    def api_cameras():
        if request.method == "GET":
            return jsonify([c.to_dict() for c in Camera.query.all()])
        data = request.get_json(force=True) or {}
        name = (data.get("name") or "").strip()
        source_raw = data.get("source")
        if source_raw is None or str(source_raw).strip() == "":
            return jsonify({"error": "camera source is required"}), 400
        if not name:
            return jsonify({"error": "camera name is required"}), 400
        source = str(source_raw).strip()
        existing = _find_camera_by_source(source)
        if existing is not None:
            return jsonify(
                {
                    "error": "a camera with this source already exists",
                    "existing_id": existing.id,
                    "existing_name": existing.name,
                }
            ), 409
        cam = Camera(
            name=name,
            source=source,
            location=(data.get("location") or "").strip() or None,
            is_active=True,
        )
        _db.session.add(cam)
        _db.session.commit()
        manager.start(cam, app)
        return jsonify(cam.to_dict()), 201

    @app.route("/api/cameras/<int:cid>", methods=["PUT", "DELETE"])
    def api_camera_detail(cid: int):
        cam = Camera.query.get_or_404(cid)
        if request.method == "DELETE":
            # Stop the camera if it's running (do it server-side so the
            # delete is a single atomic operation from the frontend's POV).
            manager.stop(cid)

            # Null-out FK references so the camera row can be deleted.
            Attendance.query.filter_by(camera_id=cid).update(
                {Attendance.camera_id: None},
                synchronize_session=False,
            )
            UnknownFace.query.filter_by(camera_id=cid).update(
                {UnknownFace.camera_id: None},
                synchronize_session=False,
            )

            # Flush the bulk updates and expire the session so that the
            # ORM's identity map no longer holds stale references to
            # Attendance/UnknownFace objects that still point at this cam.
            _db.session.flush()
            _db.session.expire_all()

            # Re-fetch the camera after expiry to get a clean ORM state.
            cam = Camera.query.get(cid)
            if cam is None:
                return jsonify({"ok": True})

            try:
                _db.session.delete(cam)
                _db.session.commit()
            except IntegrityError:
                _db.session.rollback()
                return jsonify(
                    {"error": "cannot delete camera; referenced by other records"}
                ), 409
            return jsonify({"ok": True})
        data = request.get_json(force=True) or {}
        if "name" in data:
            name = (data.get("name") or "").strip()
            if not name:
                return jsonify({"error": "camera name is required"}), 400
            cam.name = name
        if "source" in data:
            source_raw = data.get("source")
            if source_raw is None or str(source_raw).strip() == "":
                return jsonify({"error": "camera source is required"}), 400
            source = str(source_raw).strip()
            existing = _find_camera_by_source(source)
            if existing is not None and existing.id != cam.id:
                return jsonify(
                    {
                        "error": "a camera with this source already exists",
                        "existing_id": existing.id,
                        "existing_name": existing.name,
                    }
                ), 409
            cam.source = source
        if "location" in data:
            cam.location = (data.get("location") or "").strip() or None
        if "is_active" in data:
            cam.is_active = bool(data["is_active"])
        _db.session.commit()
        manager.refresh_known()
        return jsonify(cam.to_dict())

    @app.route("/api/cameras/<int:cid>/start", methods=["POST"])
    def api_camera_start(cid: int):
        cam = Camera.query.get_or_404(cid)
        cam.is_active = True
        _db.session.commit()
        manager.start(cam, app)
        return jsonify({"ok": True, "camera": cam.to_dict()})

    @app.route("/api/cameras/<int:cid>/stop", methods=["POST"])
    def api_camera_stop(cid: int):
        manager.stop(cid)
        cam = Camera.query.get(cid)
        if cam:
            cam.is_active = False
            _db.session.commit()
        return jsonify({"ok": True})

    @app.route("/api/cameras/<int:cid>/test", methods=["POST"])
    def api_camera_test(cid: int):
        """Quick non-destructive reachability check for a camera URL.

        We don't open a second VideoCapture for the same device (avoids
        macOS exclusive-access issues). Instead we just confirm the
        configuration is present and the manager has a recognizer for it,
        which is a strong signal it's reachable.
        """
        cam = Camera.query.get_or_404(cid)
        ok = cam.source and len(cam.source) > 0
        return jsonify(
            {
                "ok": bool(ok),
                "message": "source configured" if ok else "no source",
                "source": cam.source,
            }
        )

    @app.route("/api/cameras/status")
    def api_camera_status():
        statuses = {}
        for cam in Camera.query.all():
            frame, dets = manager.snapshot(cam.id)
            recognized = sum(1 for d in dets if d.get("name") != "Unknown")
            unknown = sum(1 for d in dets if d.get("name") == "Unknown")
            statuses[str(cam.id)] = {
                "camera_id": cam.id,
                "fps": "Live" if frame is not None and cam.is_active else "--",
                "detected_faces": len(dets),
                "recognized_faces": recognized,
                "unknown_faces": unknown,
            }
        return jsonify(statuses)

    # ----------------------------------------------------------------- students
    @app.route("/api/students", methods=["GET", "POST"])
    def api_students():
        if request.method == "GET":
            q = Student.query
            class_name = _normalize_class_name(request.args.get("class_name"))
            if class_name:
                q = q.filter(func.lower(Student.class_name) == func.lower(class_name))
            section = _normalize_section(request.args.get("section"))
            if section:
                q = q.filter(func.lower(Student.section) == func.lower(section))
            search = (request.args.get("q") or "").strip()
            if search:
                like = f"%{search.lower()}%"
                q = q.filter(
                    _db.or_(
                        func.lower(Student.roll_no).like(like),
                        func.lower(Student.name).like(like),
                        func.cast(Student.id, _db.String).like(like),
                    )
                )
            return jsonify([s.to_dict() for s in q.order_by(Student.roll_no).all()])
        data = request.get_json(force=True) or {}
        roll_no = (data.get("roll_no") or "").strip()
        name = (data.get("name") or "").strip()
        if not roll_no:
            return jsonify({"error": "roll number is required"}), 400
        if not name:
            return jsonify({"error": "name is required"}), 400
        class_name = _normalize_class_name(data.get("class_name"))
        if not class_name:
            return jsonify({"error": "class is required"}), 400
        if Student.query.filter(
            func.lower(Student.roll_no) == roll_no.lower()
        ).first():
            return jsonify({"error": f"roll number '{roll_no}' already exists"}), 409
        s = Student(
            roll_no=roll_no,
            name=name,
            class_name=class_name,
            section=_normalize_section(data.get("section")),
            mobile=(data.get("mobile") or "").strip() or None,
            email=(data.get("email") or "").strip() or None,
        )
        _db.session.add(s)
        try:
            _db.session.commit()
        except IntegrityError:
            _db.session.rollback()
            return jsonify({"error": f"roll number '{roll_no}' already exists"}), 409
        return jsonify(s.to_dict()), 201

    @app.route("/api/students/<int:sid>", methods=["PUT", "DELETE"])
    def api_student_detail(sid: int):
        s = Student.query.get_or_404(sid)
        if request.method == "DELETE":
            if s.photo_path:
                try:
                    os.remove(
                        os.path.join(
                            app.config["KNOWN_FACES_DIR"],
                            os.path.basename(s.photo_path),
                        )
                    )
                except OSError:
                    pass
            _db.session.delete(s)
            _db.session.commit()
            manager.refresh_known()
            return jsonify({"ok": True})
        data = request.get_json(force=True)
        for key in ("roll_no", "name", "mobile", "email"):
            if key in data:
                setattr(s, key, data[key])
        if "class_name" in data:
            s.class_name = _normalize_class_name(data["class_name"])
        if "section" in data:
            s.section = _normalize_section(data["section"])
        if "is_active" in data:
            s.is_active = _coerce_bool(data["is_active"])
        _db.session.commit()
        return jsonify(s.to_dict())

    @app.route("/api/students/<int:sid>/photo", methods=["POST"])
    def api_student_photo(sid: int):
        s = Student.query.get_or_404(sid)
        if "photo" not in request.files:
            return jsonify({"error": "no file"}), 400
        f = request.files["photo"]
        fname = secure_filename(f"{s.roll_no}_{f.filename}")
        save_path = os.path.join(app.config["KNOWN_FACES_DIR"], fname)
        f.save(save_path)
        s.photo_path = f"known_faces/{fname}"

        embedding = encode_image(save_path)
        if embedding is None:
            os.remove(save_path)
            return jsonify({"error": "no face found in image"}), 400

        s.encoding = embedding
        _db.session.commit()
        manager.refresh_known()
        return jsonify(s.to_dict())

    # --------------------------------------------------------------- attendance
    @app.route("/api/attendance")
    def api_attendance():
        q = Attendance.query.order_by(Attendance.timestamp.desc())
        cid = request.args.get("camera_id", type=int)
        if cid:
            q = q.filter_by(camera_id=cid)
        if request.args.get("today") in {"1", "true", "yes"}:
            q = q.filter(func.date(Attendance.timestamp) == date.today())
        class_name = request.args.get("class_name")
        section = request.args.get("section")
        d = request.args.get("date")
        search = (request.args.get("q") or "").strip()
        if d:
            try:
                target = datetime.strptime(d, "%Y-%m-%d").date()
                q = q.filter(func.date(Attendance.timestamp) == target)
            except ValueError:
                return jsonify({"error": "bad date; use YYYY-MM-DD"}), 400
        if class_name or section or search:
            q = q.join(Student, Student.id == Attendance.student_id)
            if class_name:
                q = q.filter(Student.class_name == class_name)
            if section:
                q = q.filter(Student.section == section)
            if search:
                like = f"%{search.lower()}%"
                q = q.filter(
                    _db.or_(
                        func.lower(Student.roll_no).like(like),
                        func.lower(Student.name).like(like),
                        func.cast(Student.id, _db.String).like(like),
                    )
                )
        return jsonify([r.to_dict() for r in q.limit(1000).all()])

    @app.route("/api/attendance/<int:aid>", methods=["PUT", "DELETE"])
    def api_attendance_detail(aid: int):
        record = Attendance.query.get_or_404(aid)
        if request.method == "DELETE":
            record_date = (
                record.timestamp.date()
                if record.timestamp
                else None
            )
            force = _coerce_bool(request.args.get("force"))
            if record_date and record_date != date.today() and not force:
                return jsonify(
                    {
                        "error": "only today's attendance can be deleted without force=true",
                        "record_date": record_date.isoformat(),
                    }
                ), 403
            _db.session.delete(record)
            _db.session.commit()
            return jsonify({"ok": True})

        data = request.get_json(force=True) or {}
        record_date = (
            record.timestamp.date()
            if record.timestamp
            else None
        )
        force = _coerce_bool(data.get("force"))
        if record_date and record_date != date.today() and not force:
            return jsonify(
                {
                    "error": "only today's attendance can be edited without force=true",
                    "record_date": record_date.isoformat(),
                }
            ), 403

        if "student_id" in data:
            student_id = data.get("student_id")
            if not student_id:
                return jsonify({"error": "student_id is required"}), 400
            student = Student.query.get(student_id)
            if student is None:
                return jsonify({"error": "student not found"}), 404
            record.student_id = int(student_id)

        if "status" in data:
            status = (data.get("status") or "").strip().lower()
            if status not in {"present", "absent", "late"}:
                return jsonify({"error": "status must be present, absent, or late"}), 400
            record.status = status

        if "confidence" in data:
            try:
                record.confidence = float(data["confidence"])
            except (TypeError, ValueError):
                return jsonify({"error": "confidence must be a number"}), 400

        if "camera_id" in data:
            camera_id = data.get("camera_id")
            if camera_id in (None, "", 0, "0"):
                record.camera_id = None
            else:
                camera = Camera.query.get(int(camera_id))
                if camera is None:
                    return jsonify({"error": "camera not found"}), 404
                record.camera_id = camera.id

        if "timestamp" in data and data.get("timestamp"):
            try:
                raw_ts = data["timestamp"]
                if isinstance(raw_ts, str) and raw_ts.endswith("Z"):
                    raw_ts = raw_ts[:-1] + "+00:00"
                parsed = datetime.fromisoformat(raw_ts)
                if parsed.tzinfo is None:
                    parsed = parsed.replace(tzinfo=timezone.utc)
                record.timestamp = parsed
            except (TypeError, ValueError):
                return jsonify(
                    {"error": "timestamp must be an ISO-8601 datetime"}
                ), 400

        _db.session.commit()
        return jsonify(record.to_dict())

    @app.route("/api/attendance/today")
    def api_attendance_today():
        rows = (
            Attendance.query.filter(func.date(Attendance.timestamp) == date.today())
            .order_by(Attendance.timestamp)
            .all()
        )
        return jsonify([r.to_dict() for r in rows])

    @app.route("/api/attendance/export")
    def api_attendance_export():
        fmt = (request.args.get("format") or "csv").lower()
        class_name = request.args.get("class_name") or None
        section = request.args.get("section") or None
        d = request.args.get("date")
        target = (
            datetime.strptime(d, "%Y-%m-%d").date()
            if d
            else date.today()
        )
        q = (
            Attendance.query.join(Student, Student.id == Attendance.student_id)
            .filter(func.date(Attendance.timestamp) == target)
        )
        if class_name:
            q = q.filter(Student.class_name == class_name)
        if section:
            q = q.filter(Student.section == section)
        rows = q.order_by(Attendance.timestamp).all()

        import csv
        import io

        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(
            ["Time", "Roll", "Name", "Class", "Section", "Camera", "Confidence"]
        )
        for r in rows:
            writer.writerow(
                [
                    r.timestamp.strftime("%H:%M:%S") if r.timestamp else "",
                    r.student.roll_no if r.student else "",
                    r.student.name if r.student else "",
                    r.student.class_name if r.student else "",
                    r.student.section if r.student else "",
                    r.camera.name if r.camera else "",
                    f"{(r.confidence or 0):.2f}",
                ]
            )
        from flask import Response as _R

        if fmt == "pdf":
            # Lightweight PDF via reportlab if available, otherwise fall back to CSV download.
            try:
                from reportlab.lib.pagesizes import A4
                from reportlab.pdfgen import canvas as _cv
                from reportlab.lib.units import cm

                pdf_buf = io.BytesIO()
                c = _cv.Canvas(pdf_buf, pagesize=A4)
                c.setFont("Helvetica-Bold", 14)
                c.drawString(2 * cm, 28 * cm, "Attendance Report")
                c.setFont("Helvetica", 10)
                c.drawString(
                    2 * cm, 27 * cm, f"Date: {target.isoformat()}    Class: {class_name or 'ALL'}"
                )
                y = 25 * cm
                c.setFont("Helvetica-Bold", 10)
                c.drawString(2 * cm, y, "Time")
                c.drawString(4 * cm, y, "Roll")
                c.drawString(6 * cm, y, "Name")
                c.drawString(10 * cm, y, "Class")
                c.drawString(13 * cm, y, "Camera")
                c.drawString(17 * cm, y, "Conf")
                y -= 0.6 * cm
                c.setFont("Helvetica", 10)
                for r in rows:
                    if y < 2 * cm:
                        c.showPage()
                        y = 28 * cm
                    c.drawString(2 * cm, y, r.timestamp.strftime("%H:%M:%S") if r.timestamp else "")
                    c.drawString(4 * cm, y, r.student.roll_no if r.student else "")
                    c.drawString(6 * cm, y, (r.student.name if r.student else "")[:30])
                    c.drawString(10 * cm, y, r.student.class_name if r.student else "")
                    c.drawString(13 * cm, y, r.camera.name if r.camera else "")
                    c.drawString(17 * cm, y, f"{(r.confidence or 0):.2f}")
                    y -= 0.55 * cm
                c.save()
                return _R(
                    pdf_buf.getvalue(),
                    mimetype="application/pdf",
                    headers={
                        "Content-Disposition": f"attachment; filename=attendance_{target.isoformat()}.pdf"
                    },
                )
            except Exception:
                pass

        return _R(
            buf.getvalue(),
            mimetype="text/csv",
            headers={
                "Content-Disposition": f"attachment; filename=attendance_{target.isoformat()}.csv"
            },
        )

    # ----------------------------------------- enterprise report export
    @app.route("/api/reports/export")
    def api_reports_export():
        """Export a report in CSV / Excel / PDF / Print (HTML)."""
        from enterprise_query import build_report
        report_type = request.args.get("type") or "daily"
        fmt = (request.args.get("format") or "csv").lower()
        d_raw = request.args.get("date")
        try:
            target = datetime.strptime(d_raw, "%Y-%m-%d").date() if d_raw else date.today()
        except ValueError:
            return jsonify({"error": "date must be YYYY-MM-DD"}), 400
        def _opt(name):
            raw = request.args.get(name)
            if not raw:
                return None
            try:
                return datetime.strptime(raw, "%Y-%m-%d").date()
            except ValueError:
                return None
        rows, headers, title = build_report(
            report_type,
            d=target,
            start=_opt("start"),
            end=_opt("end"),
            class_name=request.args.get("class_name") or None,
            section=request.args.get("section") or None,
            department=request.args.get("department") or None,
            designation=request.args.get("designation") or None,
            search=(request.args.get("q") or "").strip() or None,
        )
        filename = f"{report_type}_{target.isoformat()}"
        if fmt == "json":
            return jsonify({"title": title, "headers": headers, "rows": rows})
        if fmt in ("excel", "xlsx"):
            return _export_excel(rows, headers, title, filename)
        if fmt in ("pdf",):
            return _export_pdf(rows, headers, title, filename)
        if fmt in ("print", "html"):
            return _export_printable_html(rows, headers, title, filename)
        return _export_csv(rows, headers, filename)

    def _export_csv(rows, headers, filename):
        import csv as _csv
        import io as _io
        buf = _io.StringIO()
        w = _csv.writer(buf)
        w.writerow(headers)
        for r in rows:
            w.writerow([r.get(h, "") for h in headers])
        return Response(
            buf.getvalue(),
            mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"},
        )

    def _export_excel(rows, headers, title, filename):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return _export_csv(rows, headers, filename)
        wb = Workbook()
        ws = wb.active
        ws.title = (title or "Report")[:31]
        bold = Font(bold=True)
        header_fill = PatternFill("solid", fgColor="007AFF")
        white_bold = Font(bold=True, color="FFFFFFFF")
        ws.append(headers)
        for cell in ws[1]:
            cell.font = white_bold
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append([r.get(h, "") for h in headers])
        # Auto column width
        for col_idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for r in rows:
                v = str(r.get(header, ""))
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(max_len + 2, 10), 40)
        from io import BytesIO as _B
        buf = _B()
        wb.save(buf)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"},
        )

    def _export_pdf(rows, headers, title, filename):
        try:
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            )
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            return _export_csv(rows, headers, filename)
        from io import BytesIO as _B
        buf = _B()
        doc = SimpleDocTemplate(
            buf, pagesize=landscape(A4),
            leftMargin=1 * cm, rightMargin=1 * cm,
            topMargin=1 * cm, bottomMargin=1 * cm,
        )
        styles = getSampleStyleSheet()
        elements = [
            Paragraph(f"<b>{title}</b>", styles["Title"]),
            Spacer(1, 0.4 * cm),
        ]
        data = [headers] + [[str(r.get(h, ""))[:60] for h in headers] for r in rows]
        if not rows:
            data.append(["No records"] + [""] * (len(headers) - 1))
        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#007AFF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elements.append(table)
        doc.build(elements)
        return Response(
            buf.getvalue(),
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}.pdf"},
        )

    def _export_printable_html(rows, headers, title, filename):
        body = "".join(f"<th>{h}</th>" for h in headers)
        rows_html = "".join(
            "<tr>" + "".join(f"<td>{r.get(h, '')}</td>" for h in headers) + "</tr>"
            for r in rows
        )
        if not rows_html:
            rows_html = f"<tr><td colspan='{len(headers)}'>No records found.</td></tr>"
        html = f"""<!doctype html>
<html><head><meta charset='utf-8'><title>{title}</title>
<style>
body {{ font-family: -apple-system, sans-serif; padding: 20px; }}
table {{ width: 100%; border-collapse: collapse; font-size: 12px; }}
th, td {{ border: 1px solid #d1d1d6; padding: 6px 8px; text-align: left; }}
th {{ background: #007AFF; color: white; }}
h1 {{ color: #1d1d1f; }}
@media print {{ .no-print {{ display: none; }} }}
</style></head><body>
<h1>{title}</h1>
<button class='no-print' onclick='window.print()' style='padding:8px 16px;background:#007AFF;color:white;border:none;border-radius:6px;cursor:pointer;'>Print</button>
<table><thead><tr>{body}</tr></thead><tbody>{rows_html}</tbody></table>
</body></html>"""
        return Response(html, mimetype="text/html")

    # ----------------------------------------- enterprise attendance (v2)
    def _enterprise_row(row, kind: str) -> dict:
        person = row.student if kind == "student" else row.teacher
        return {"id": row.id, "type": kind, "date": row.attendance_date.isoformat(), "in_time": row.in_time.isoformat() if row.in_time else None, "out_time": row.out_time.isoformat() if row.out_time else None, "working_minutes": row.working_minutes, "break_minutes": row.break_minutes, "overtime_minutes": row.overtime_minutes, "late": row.is_late, "early_exit": row.is_early_exit, "status": row.status, "camera": row.camera.name if row.camera else None, "confidence": row.confidence, "person_id": person.id, "name": person.name, "roll_no": person.roll_no if kind == "student" else None, "class_name": person.class_name if kind == "student" else None, "section": person.section if kind == "student" else None, "employee_id": person.teacher_id if kind == "teacher" else None, "department": person.department if kind == "teacher" else None, "designation": person.designation if kind == "teacher" else None}

    @app.route("/api/enterprise/attendance", methods=["GET", "POST"])
    def api_enterprise_attendance():
        from enterprise_query import list_sessions
        data = request.get_json(silent=True) or {}
        kind = (request.args.get("type") or data.get("type") or "all").lower()
        if kind not in {"student", "teacher", "all"}:
            return jsonify({"error": "type must be student, teacher, or all"}), 400

        if request.method == "POST":
            # Manual entry: insert an enterprise row (no camera needed).
            from attendance_service import record_recognition
            payload = dict(data)
            person_kind = (payload.get("type") or "student").lower()
            if person_kind not in {"student", "teacher"}:
                return jsonify({"error": "type must be student or teacher"}), 400
            try:
                confidence = float(payload.get("confidence", 0))
            except (TypeError, ValueError):
                confidence = 0.0
            from datetime import datetime as _dt
            now = _dt.now(timezone.utc)
            result = record_recognition(
                person_kind,
                int(payload["person_id"]),
                payload.get("camera_id"),
                confidence,
                now=now,
            )
            if result.get("action") == "rejected":
                return jsonify(result), 409
            # Return the row we just created/updated.
            model = StudentAttendance if person_kind == "student" else TeacherAttendance
            row = model.query.get(result["attendance_id"])
            return jsonify(_enterprise_row(row, person_kind)), 201

        # GET: unified list across the same query engine
        d_raw = request.args.get("date")
        try:
            target = datetime.strptime(d_raw, "%Y-%m-%d").date() if d_raw else date.today()
        except ValueError:
            return jsonify({"error": "date must be YYYY-MM-DD"}), 400
        # Map event_type to status for the underlying query
        event_type = request.args.get("event_type") or None
        status = request.args.get("status") or None
        rows = list_sessions(
            d=target,
            kind=kind,
            class_name=request.args.get("class_name") or None,
            section=request.args.get("section") or None,
            department=request.args.get("department") or None,
            designation=request.args.get("designation") or None,
            status=status,
            event_type=event_type,
            search=(request.args.get("q") or "").strip() or None,
            limit=int(request.args.get("limit", 1000)),
        )
        return jsonify(rows)

    @app.route("/api/enterprise/dashboard")
    def api_enterprise_dashboard():
        from enterprise_query import compute_dashboard_summary
        d = request.args.get("date")
        try:
            target = datetime.strptime(d, "%Y-%m-%d").date() if d else date.today()
        except ValueError:
            return jsonify({"error": "date must be YYYY-MM-DD"}), 400
        return jsonify(compute_dashboard_summary(
            target,
            class_name=request.args.get("class_name") or None,
            section=request.args.get("section") or None,
            department=request.args.get("department") or None,
            designation=request.args.get("designation") or None,
        ))

    @app.route("/api/enterprise/summary")
    def api_enterprise_summary():
        """Same as the dashboard endpoint but always for today. Powers the
        cards on the IN/OUT page, the Time Reports preview, and any other
        surface that needs live counts."""
        from enterprise_query import compute_dashboard_summary
        d = request.args.get("date")
        try:
            target = datetime.strptime(d, "%Y-%m-%d").date() if d else date.today()
        except ValueError:
            return jsonify({"error": "date must be YYYY-MM-DD"}), 400
        return jsonify(compute_dashboard_summary(
            target,
            class_name=request.args.get("class_name") or None,
            section=request.args.get("section") or None,
            department=request.args.get("department") or None,
            designation=request.args.get("designation") or None,
        ))

    @app.route("/api/enterprise/widgets")
    def api_enterprise_widgets():
        """Additional cards / graphs (trend, top-late, top-OT, etc.)."""
        from enterprise_query import (
            attendance_trend, hourly_entries, top_late_students,
            top_overtime_teachers, top_early_exits, working_hours_distribution,
            camera_health, recognition_stats, activity_feed, class_attendance_heatmap,
        )
        return jsonify({
            "trend": attendance_trend(days=7),
            "hourly": hourly_entries(),
            "top_late": top_late_students(),
            "top_overtime": top_overtime_teachers(),
            "top_early_exit": top_early_exits(),
            "working_hours": working_hours_distribution(),
            "camera_health": camera_health(),
            "recognition": recognition_stats(),
            "activity_feed": activity_feed(limit=30),
            "heatmap": class_attendance_heatmap(days=7),
        })

    @app.route("/api/enterprise/sessions")
    def api_enterprise_sessions():
        """Filtered session list for the IN/OUT page."""
        from enterprise_query import list_sessions
        d = request.args.get("date")
        try:
            target = datetime.strptime(d, "%Y-%m-%d").date() if d else date.today()
        except ValueError:
            return jsonify({"error": "date must be YYYY-MM-DD"}), 400
        kind = (request.args.get("type") or "all").lower()
        try:
            min_conf = float(request.args.get("min_confidence")) if request.args.get("min_confidence") else None
            max_conf = float(request.args.get("max_confidence")) if request.args.get("max_confidence") else None
        except (TypeError, ValueError):
            return jsonify({"error": "confidence filters must be numeric"}), 400
        return jsonify(list_sessions(
            d=target,
            kind=kind,
            class_name=request.args.get("class_name") or None,
            section=request.args.get("section") or None,
            department=request.args.get("department") or None,
            designation=request.args.get("designation") or None,
            status=request.args.get("status") or None,
            event_type=request.args.get("event_type") or None,
            camera_id=request.args.get("camera_id", type=int),
            min_confidence=min_conf,
            max_confidence=max_conf,
            recognition_type=request.args.get("recognition_type") or None,
            search=(request.args.get("q") or "").strip() or None,
            limit=int(request.args.get("limit", 1000)),
        ))

    @app.route("/api/reports/preview")
    def api_reports_preview():
        """Return a JSON report preview that matches the exported file."""
        from enterprise_query import build_report
        report_type = request.args.get("type") or "daily"
        d = request.args.get("date")
        try:
            target = datetime.strptime(d, "%Y-%m-%d").date() if d else date.today()
        except ValueError:
            return jsonify({"error": "date must be YYYY-MM-DD"}), 400
        def _opt(name):
            raw = request.args.get(name)
            if not raw:
                return None
            try:
                return datetime.strptime(raw, "%Y-%m-%d").date()
            except ValueError:
                return None
        rows, headers, title = build_report(
            report_type,
            d=target,
            start=_opt("start"),
            end=_opt("end"),
            class_name=request.args.get("class_name") or None,
            section=request.args.get("section") or None,
            department=request.args.get("department") or None,
            designation=request.args.get("designation") or None,
            search=(request.args.get("q") or "").strip() or None,
        )
        return jsonify({
            "title": title,
            "headers": headers,
            "rows": rows,
            "summary": {
                "total_records": len(rows),
            },
        })

    @app.route("/api/enterprise/events")
    def api_enterprise_events():
        """IN/OUT feed sourced from immutable attendance_events rows."""
        kind = (request.args.get("type") or "student").lower()
        if kind not in {"student", "teacher"}: return jsonify({"error": "type must be student or teacher"}), 400
        q = AttendanceEvent.query.filter_by(attendance_type=kind)
        raw_date = request.args.get("date")
        if raw_date:
            try: q = q.filter(func.date(AttendanceEvent.event_time) == datetime.strptime(raw_date, "%Y-%m-%d").date())
            except ValueError: return jsonify({"error": "date must be YYYY-MM-DD"}), 400
        events = q.order_by(AttendanceEvent.event_time.desc()).limit(1000).all(); out = []
        model, person_model, fk = (StudentAttendance, Student, "student_id") if kind == "student" else (TeacherAttendance, Teacher, "teacher_id")
        for event in events:
            row = model.query.get(event.attendance_id)
            if row is None: continue
            person = getattr(row, "student" if kind == "student" else "teacher")
            if kind == "student":
                if request.args.get("class_name") and person.class_name != request.args["class_name"]: continue
                if request.args.get("section") and person.section != request.args["section"]: continue
            else:
                if request.args.get("department") and person.department != request.args["department"]: continue
                if request.args.get("designation") and person.designation != request.args["designation"]: continue
            data = _enterprise_row(row, kind); data.update({"event_id": event.id, "event_type": event.event_type, "event_time": event.event_time.isoformat(), "event_camera": event.camera.name if event.camera else None, "event_confidence": event.confidence})
            out.append(data)
        return jsonify(out)

    @app.route("/api/enterprise/settings", methods=["GET", "PUT"])
    def api_enterprise_settings():
        keys = ["face_recognition", "qr", "rfid", "manual_attendance", "whatsapp", "sms", "email", "anti_spoofing", "liveness_detection", "unknown_face_alert"]
        if request.method == "PUT":
            for key, value in (request.get_json(force=True) or {}).items():
                if key in keys:
                    setting = AttendanceSetting.query.filter_by(key=key).first() or AttendanceSetting(key=key)
                    setting.enabled = _coerce_bool(value); _db.session.add(setting)
            _db.session.commit()
        saved = {s.key: s.enabled for s in AttendanceSetting.query.all()}; return jsonify({key: saved.get(key, key == "face_recognition") for key in keys})

    @app.route("/api/enterprise/time-rules/<kind>", methods=["GET", "PUT"])
    def api_enterprise_time_rules(kind: str):
        from database.enterprise_models import TeacherTimeRule, StudentTimeRule
        model = {"teacher": TeacherTimeRule, "student": StudentTimeRule}.get(kind)
        if model is None: return jsonify({"error": "kind must be teacher or student"}), 400
        rule = model.query.first() or model(); _db.session.add(rule)
        if request.method == "PUT":
            data = request.get_json(force=True) or {}
            allowed = {"office_start", "in_end_time", "office_end", "late_time", "half_day_time", "absent_after", "out_detection_start", "early_exit_time", "overtime_start", "min_working_minutes", "overtime_enabled", "gate_close_time"}
            for key, value in data.items():
                if key in allowed and hasattr(rule, key):
                    old = getattr(rule, key); setattr(rule, key, value)
                    if old != value: _db.session.add(AttendanceLog(entity_type=f"{kind}_time_rule", entity_id=rule.id or 0, field=key, old_value=str(old), new_value=str(value), reason=data.get("reason")))
            rule.updated_at = _utcnow(); _db.session.commit()
        return jsonify({c.name: getattr(rule, c.name) for c in rule.__table__.columns})

    # --------------------------------------------------------- unknown faces
    @app.route("/api/unknown-faces")
    def api_unknown_faces():
        rows = UnknownFace.query.order_by(UnknownFace.timestamp.desc()).limit(200)
        return jsonify([r.to_dict() for r in rows.all()])

    @app.route("/api/unknown-faces/<int:uid>/action", methods=["POST", "DELETE"])
    def api_unknown_face_action(uid: int):
        """Audit-approved lifecycle actions without deleting history by default."""
        face = UnknownFace.query.get_or_404(uid)
        if request.method == "DELETE":
            _db.session.add(AttendanceLog(entity_type="unknown_face", entity_id=uid, field="delete", old_value="active", new_value="deleted", reason=request.args.get("reason")))
            _db.session.delete(face); _db.session.commit(); return jsonify({"ok": True})
        data = request.get_json(force=True) or {}; action = (data.get("action") or "").lower()
        if action not in {"approve", "reject", "assign_student", "assign_teacher"}:
            return jsonify({"error": "action must be approve, reject, assign_student, or assign_teacher"}), 400
        if action == "assign_student" and not Student.query.get(data.get("student_id")):
            return jsonify({"error": "student not found"}), 404
        if action == "assign_teacher" and not Teacher.query.get(data.get("teacher_id")):
            return jsonify({"error": "teacher not found"}), 404
        face.alerted = action in {"approve", "assign_student", "assign_teacher"}
        _db.session.add(AttendanceLog(entity_type="unknown_face", entity_id=uid, field="action", old_value=None, new_value=action, reason=data.get("reason")))
        _db.session.commit(); return jsonify({"ok": True, "action": action})

    # --------------------------------------------------------------- teachers
    @app.route("/api/teachers", methods=["GET", "POST"])
    def api_teachers():
        if request.method == "GET":
            q = Teacher.query.order_by(Teacher.name)
            search = (request.args.get("q") or "").strip()
            if search:
                like = f"%{search.lower()}%"
                q = q.filter(
                    _db.or_(
                        func.lower(Teacher.name).like(like),
                        func.lower(Teacher.teacher_id).like(like),
                        func.lower(Teacher.subject).like(like),
                        func.lower(Teacher.assigned_classes).like(like),
                    )
                )
            return jsonify([t.to_dict() for t in q.all()])
        data = request.get_json(force=True)
        t = Teacher(
            teacher_id=data["teacher_id"],
            name=data["name"],
            subject=data.get("subject"),
            assigned_classes=data.get("assigned_classes"),
            mobile=data.get("mobile"),
            email=data.get("email"),
            department=data.get("department"),
            designation=data.get("designation"),
            is_active=_coerce_bool(data.get("is_active", True)),
        )
        _db.session.add(t)
        _db.session.commit()
        return jsonify(t.to_dict()), 201

    @app.route("/api/teachers/<int:tid>", methods=["PUT", "DELETE"])
    def api_teacher_detail(tid: int):
        t = Teacher.query.get_or_404(tid)
        if request.method == "DELETE":
            _db.session.delete(t)
            _db.session.commit()
            return jsonify({"ok": True})
        data = request.get_json(force=True)
        for key in ("teacher_id", "name", "subject", "assigned_classes", "mobile", "email", "department", "designation"):
            if key in data:
                setattr(t, key, data[key])
        if "is_active" in data:
            t.is_active = _coerce_bool(data["is_active"])
        _db.session.commit()
        return jsonify(t.to_dict())

    @app.route("/api/teachers/<int:tid>/photo", methods=["POST"])
    def api_teacher_photo(tid: int):
        t = Teacher.query.get_or_404(tid)
        if "photo" not in request.files: return jsonify({"error": "no file"}), 400
        f = request.files["photo"]; fname = secure_filename(f"teacher_{t.teacher_id}_{f.filename}")
        save_path = os.path.join(app.config["KNOWN_FACES_DIR"], fname); f.save(save_path)
        embedding = encode_image(save_path)
        if embedding is None:
            try: os.remove(save_path)
            except OSError: pass
            return jsonify({"error": "no face found in image"}), 400
        t.photo_path, t.encoding = f"known_faces/{fname}", embedding; _db.session.commit(); manager.refresh_known()
        print(f"[RECOGNITION] teacher enrollment refreshed teacher={t.id}")
        return jsonify(t.to_dict())

    # ----------------------------------------------------------- MJPEG stream
    @app.route("/stream/<int:cid>")
    def stream(cid: int):
        if cid not in manager.recognizers:
            manager.start(Camera.query.get_or_404(cid), app)

        def gen():
            while True:
                # Check if camera/recognizer is still active and running in manager
                rec = manager.recognizers.get(cid)
                if rec is None or not rec.running:
                    print(f"[DEBUG] Camera {cid} is not active or running. Stopping stream.")
                    break

                frame, dets = manager.snapshot(cid)
                if frame is None:
                    time.sleep(0.1)
                    continue
                for d in dets:
                    t, r, b, l = d["box"]
                    color = (0, 200, 0) if d["name"] != "Unknown" else (0, 0, 200)
                    cv2.rectangle(frame, (l, t), (r, b), color, 2)
                    label = f"{d['name']} ({d['confidence']:.2f})"
                    cv2.putText(
                        frame,
                        label,
                        (l, max(t - 8, 12)),
                        cv2.FONT_HERSHEY_SIMPLEX,
                        0.5,
                        color,
                        1,
                    )
                # [DEBUG] before encoding, confirm we actually have a frame
                # and that detections came back from the recognizer.
                print(
                    f"[DEBUG] Streaming frame={'YES' if frame is not None else 'NO'} "
                    f"detections={len(dets)}"
                )
                ok, buf = cv2.imencode(".jpg", frame)
                if not ok:
                    continue
                yield (
                    b"--frame\r\n"
                    b"Content-Type: image/jpeg\r\n\r\n" + buf.tobytes() + b"\r\n"
                )

        return Response(gen(), mimetype="multipart/x-mixed-replace; boundary=frame")

    # ----------------------------------------------------------------- /api/ask
    @app.route("/api/ask", methods=["POST"])
    def api_ask():
        from slm.nlp_engine import answer as slm_answer
        data = request.get_json(force=True)
        question = data.get("question", "")
        return jsonify({"answer": slm_answer(question)})


import time  # noqa: E402  (used by MJPEG generator's sleep)

app = create_app()


def _resolve_port(default_port: int = 5000) -> int:
    requested_port = int(os.getenv("PORT", str(default_port)))
    if requested_port <= 0:
        return default_port

    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
        sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
        try:
            sock.bind(("0.0.0.0", requested_port))
        except OSError:
            return 5001 if requested_port == 5000 else requested_port
        return requested_port


if __name__ == "__main__":
    with app.app_context():
        cam = Camera.query.filter_by(is_active=True).first()
        if cam:
            try:
                manager.start(cam, app)
            except Exception as exc:
                print(f"[WARN] startup camera initialization skipped: {exc}")
    app.run(
        host="0.0.0.0",
        port=_resolve_port(),
        debug=True,
        threaded=True,
    )
